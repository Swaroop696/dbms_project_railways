# import openpyxl
# wb = openpyxl.Workbook()
# sheet=wb.active
# sheet['A1']='Tall row'
# sheet['B2']='wide column'
# sheet.row_dimensions[1].height=70
# sheet.column_dimensions['B'].width=20
# wb.save('dimensions.xlsx')

# cells can be merged with a single cell using merge_cell
# also u can unmerge the cell using unmerge_cells

import openpyxl
wb = openpyxl.Workbook()
sheet=wb.active
sheet. merge_cells('A1:G1')
wb.save('dimensions1.xlsx')




